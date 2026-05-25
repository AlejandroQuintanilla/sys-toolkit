"""
report_generator.py - Generador de informes ejecutivos en Excel.
Módulo 7: Exporta el inventario filtrado a .xlsx con formato profesional.

Puede ejecutarse manualmente o de forma automática cada mes via schedule.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from inventory_manager import (
    load_inventory,
    filter_vulnerable,
    group_by_department,
)


def generate_excel_report(
    csv_path: str = "inventory.csv",
    output_dir: str = "output",
) -> str:
    """
    Lee el CSV más reciente, aplica filtros y genera un Excel ejecutivo.
    Devuelve la ruta del archivo generado.
    """
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m")
    filename = output_path / f"informe_servidores_{timestamp}.xlsx"

    # Carga y filtra
    df_full = load_inventory(csv_path)
    df_vulnerable = filter_vulnerable(df_full)
    df_departments = group_by_department(df_full)
    df_windows = df_full[
        df_full["os"].str.contains("Windows Server", case=False, na=False)
    ].copy()
    df_low_ram = df_full[df_full["ram_gb"] <= 4].copy()

    # Exportar a Excel con múltiples hojas
    with pd.ExcelWriter(str(filename), engine="openpyxl") as writer:

        # Hoja 1: Resumen ejecutivo por departamento
        df_departments.to_excel(
            writer, sheet_name="Resumen por Departamento", index=False
        )

        # Hoja 2: Servidores vulnerables u obsoletos
        cols_export = [
            "hostname", "ip", "os", "ram_gb", "cpu_cores",
            "department", "status", "responsible", "last_update",
        ]
        df_vulnerable[cols_export].to_excel(
            writer, sheet_name="Servidores Vulnerables", index=False
        )

        # Hoja 3: Solo Windows
        df_windows[cols_export].to_excel(
            writer, sheet_name="Windows Server", index=False
        )

        # Hoja 4: Baja RAM
        df_low_ram[cols_export].to_excel(
            writer, sheet_name="Baja RAM (<=4GB)", index=False
        )

        # Hoja 5: Inventario completo
        df_full.to_excel(
            writer, sheet_name="Inventario Completo", index=False
        )

        # Aplicar formato básico a todas las hojas
        workbook = writer.book
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        HEADER_FILL = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
        ALERT_FILL = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )

        for sheet_name in writer.sheets:
            ws = workbook[sheet_name]

            # Cabecera azul oscuro
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustar ancho de columnas automáticamente
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = max(
                    (len(str(cell.value)) if cell.value else 0 for cell in col),
                    default=10,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(
                    max_len + 3, 40
                )

            # En la hoja de vulnerables, marcar en rojo los obsoletos
            if sheet_name == "Servidores Vulnerables":
                status_col = None
                for i, cell in enumerate(ws[1], 1):
                    if cell.value == "status":
                        status_col = i
                        break
                if status_col:
                    for row in ws.iter_rows(min_row=2):
                        if row[status_col - 1].value == "obsoleto":
                            for cell in row:
                                cell.fill = ALERT_FILL

            # Congelar primera fila
            ws.freeze_panes = "A2"

    print(f"[OK] Informe Excel generado: {filename}")
    return str(filename)


def schedule_monthly_report() -> None:
    """
    Configura el módulo schedule para regenerar el informe cada mes.
    Ejecuta este script como demonio: python report_generator.py
    """
    import schedule
    import time

    print("[*] Planificador mensual de informes activo. Ctrl+C para detener.")
    # Ejecutar el primer informe inmediatamente
    generate_excel_report()

    # Programar para el día 1 de cada mes a las 07:00
    schedule.every().day.at("07:00").do(
        lambda: datetime.now().day == 1 and generate_excel_report()
    )

    while True:
        schedule.run_pending()
        time.sleep(3600)


if __name__ == "__main__":
    import sys
    if "--schedule" in sys.argv:
        schedule_monthly_report()
    else:
        generate_excel_report()
